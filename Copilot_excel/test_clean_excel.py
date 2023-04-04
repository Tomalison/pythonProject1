# 撰寫一個函式以返回特定目錄中的所有Excel檔案
import os
def get_all_excels(dir_path):
    # 建立一個空的列表，用來存放Excel檔案的絕對路徑
    excel_files = []
    # 遍歷目錄dir_path中的所有檔案
    for root, dirs, files in os.walk(dir_path):
        # 遍歷files中的所有檔案名稱
        for file in files:
            # 如果檔案名稱的副檔名為xlsx
            if file.endswith('.xlsx'):
                # 將檔案的絕對路徑加入excel_files列表中
                excel_files.append(os.path.join(root, file))
    # 返回excel_files列表
    return excel_files

path = "C:/Users/USER/PycharmProjects/pythonProject1/Copilot_excel/data"

# 呼叫get_all_excels函式以獲得特定目錄中的所有Excel檔案
excels = get_all_excels(path)
print(excels)

# 撰寫一個函式以讀取特定Excel檔案的所有工作表
import openpyxl
def get_all_worksheets(excel_file):
    # 建立一個空的列表，用來存放Excel檔案中的工作表
    all_worksheets = []
    # 開啟Excel檔案
    workbook = openpyxl.load_workbook(excel_file)
    # 遍歷Excel檔案中的所有工作表
    for sheet in workbook:
        # 將工作表名稱加入all_worksheets列表中
        all_worksheets.append(sheet.title)
    # 返回all_worksheets列表
    return all_worksheets

# 呼叫get_all_worksheets函式以獲得特定Excel檔案中的所有工作表
worksheets = get_all_worksheets(excels[0])
print(worksheets)

# 撰寫一個函式以讀取特定工作表中的所有儲存格
def get_all_cells(worksheet):
    # 建立一個空的列表，用來存放工作表中的所有儲存格
    all_cells = []
    # 遍歷工作表中的所有儲存格
    for row in worksheet.rows:
        for cell in row:
            # 將儲存格的值加入all_cells列表中
            all_cells.append(cell.value)
    # 返回all_cells列表
    return all_cells

# 呼叫get_all_cells函式以獲得特定工作表中的所有儲存格
workbook = openpyxl.load_workbook(excels[0])
worksheet = workbook[worksheets[0]]
cells = get_all_cells(worksheet)
print(cells)

# 將儲存格不為數值的儲存格改為數值0
def clean_cells(cells):
    # 建立一個空的列表，用來存放處理後的儲存格
    cleaned_cells = []
    # 遍歷cells中的所有儲存格
    for cell in cells:
        # 如果儲存格的值不為數值
        if not isinstance(cell, (int, float)):
            # 將儲存格的值改為0
            cell = 0
        # 將儲存格的值加入cleaned_cells列表中
        cleaned_cells.append(cell)
    # 返回cleaned_cells列表
    return cleaned_cells

# 呼叫clean_cells函式以清理儲存格
cleaned_cells = clean_cells(cells)
print(cleaned_cells)

# 撰寫一個函式以將清理後的儲存格寫入特定工作表中
def write_cells(worksheet, cells):
    # 遍歷cells中的所有儲存格
    for row in worksheet.rows:
        for cell in row:
            # 將儲存格的值改為cells中的值
            cell.value = cells.pop(0)
    # 將工作表中的儲存格寫入Excel檔案中
    workbook.save(excels[0])

# 呼叫write_cells函式以將清理後的儲存格寫入特定工作表中
write_cells(worksheet, cleaned_cells)

# 撰寫一個函式以清理特定Excel檔案中的所有工作表
def clean_excel(excel_file):
    # 呼叫get_all_worksheets函式以獲得特定Excel檔案中的所有工作表
    worksheets = get_all_worksheets(excel_file)
    # 遍歷worksheets中的所有工作表名稱
    for worksheet in worksheets:
        # 開啟Excel檔案
        workbook = openpyxl.load_workbook(excel_file)
        # 選擇工作表
        worksheet = workbook[worksheet]
        # 呼叫get_all_cells函式以獲得特定工作表中的所有儲存格
        cells = get_all_cells(worksheet)
        # 呼叫clean_cells函式以清理儲存格
        cleaned_cells = clean_cells(cells)
        # 呼叫write_cells函式以將清理後的儲存格寫入特定工作表中
        write_cells(worksheet, cleaned_cells)

# 呼叫clean_excel函式以清理特定Excel檔案中的所有工作表
clean_excel(excels[0])

# 撰寫一個函式以清理特定目錄中的所有Excel檔案
def clean_excels(dir_path):
    # 呼叫get_all_excels函式以獲得特定目錄中的所有Excel檔案
    excels = get_all_excels(dir_path)
    # 遍歷excels中的所有Excel檔案
    for excel in excels:
        # 呼叫clean_excel函式以清理特定Excel檔案中的所有工作表
        clean_excel(excel)

# 呼叫clean_excels函式以清理特定目錄中的所有Excel檔案
clean_excels(path)

# Path: Copilot_excel\test_clean_excel.py
# Compare this snippet from main

